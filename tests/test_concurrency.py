"""
test_concurrency.py

Benchmark Concurrent Dense Retrieval.

Measures:
- Concurrent Queries
- Successful Queries
- Failed Queries
- Total Time
- Average Latency
- Minimum Latency
- Maximum Latency
- Throughput
"""

import time
from concurrent.futures import ThreadPoolExecutor, as_completed

from app.retrieval.dense_retriever import DenseRetriever

# ==========================================================
# Configuration
# ==========================================================

QUERIES = [
    "what is fastapi",
    "what is APIRouter",
    "how does dependency injection work",
    "how is include_router implemented",
    "what is BaseRoute",
    "how are middleware added",
    "how does authentication work",
    "where are response models handled",
    "how are request objects created",
    "how are exceptions handled",
    "where is OpenAPI generated",
    "how does startup event work",
    "how are websocket routes implemented",
    "how does file upload work",
    "where is CORSMiddleware implemented",
    "how does BackgroundTasks work",
    "how are dependencies resolved",
    "where is JSONResponse implemented",
    "how are routes registered",
    "what is the lifespan context manager",
]

CONCURRENCY_LEVELS = [
    1,
    5,
    10,
    25,
    50,
    100,
]

TOP_K = 5

retriever = DenseRetriever()

# ==========================================================


def execute_search(query):

    start = time.perf_counter()

    try:
        retriever.search(
            query=query,
            top_k=TOP_K,
        )

        latency = time.perf_counter() - start

        return True, latency

    except Exception:

        latency = time.perf_counter() - start

        return False, latency


# ==========================================================

print("=" * 100)
print("DENSE RETRIEVAL CONCURRENCY BENCHMARK")
print("=" * 100)

results = []

# ==========================================================

for concurrent_queries in CONCURRENCY_LEVELS:

    print()
    print("=" * 100)
    print(f"Concurrent Queries : {concurrent_queries}")
    print("=" * 100)

    benchmark_start = time.perf_counter()

    successful = 0
    failed = 0

    latencies = []

    with ThreadPoolExecutor(max_workers=concurrent_queries) as executor:

        futures = [
            executor.submit(
                execute_search,
                QUERIES[i % len(QUERIES)],
            )
            for i in range(concurrent_queries)
        ]

        for future in as_completed(futures):

            ok, latency = future.result()

            latencies.append(latency)

            if ok:
                successful += 1
            else:
                failed += 1

    total_time = time.perf_counter() - benchmark_start

    avg_latency = (sum(latencies) / len(latencies)) * 1000
    min_latency = min(latencies) * 1000
    max_latency = max(latencies) * 1000

    throughput = successful / total_time

    results.append(
        (
            concurrent_queries,
            successful,
            failed,
            total_time,
            avg_latency,
            min_latency,
            max_latency,
            throughput,
        )
    )

    print(f"Successful Queries : {successful}")
    print(f"Failed Queries     : {failed}")
    print(f"Total Time         : {total_time:.2f} sec")
    print(f"Average Latency    : {avg_latency:.2f} ms")
    print(f"Minimum Latency    : {min_latency:.2f} ms")
    print(f"Maximum Latency    : {max_latency:.2f} ms")
    print(f"Throughput         : {throughput:.2f} queries/sec")

# ==========================================================
# Final Report
# ==========================================================

print()
print("=" * 140)
print("CONCURRENCY MATRIX")
print("=" * 140)

print(
    f"{'Concurrent Queries':<22}"
    f"{'Success':<10}"
    f"{'Failed':<10}"
    f"{'Time(s)':<12}"
    f"{'Avg(ms)':<12}"
    f"{'Min(ms)':<12}"
    f"{'Max(ms)':<12}"
    f"{'Throughput':<15}"
)

print("-" * 140)

for row in results:

    (
        concurrent_queries,
        success,
        failed,
        total,
        avg_latency,
        min_latency,
        max_latency,
        throughput,
    ) = row

    print(
        f"{concurrent_queries:<22}"
        f"{success:<10}"
        f"{failed:<10}"
        f"{total:<12.2f}"
        f"{avg_latency:<12.2f}"
        f"{min_latency:<12.2f}"
        f"{max_latency:<12.2f}"
        f"{throughput:<15.2f}"
    )

print("=" * 140)

# ==========================================================
# Export Results to Excel
# ==========================================================

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REPOSITORY_NAME = "fastapi-master"
RETRIEVER_NAME = "DenseRetriever"

excel_file = "benchmark_results.xlsx"

# ----------------------------------------------------------
# Open existing workbook or create a new one
# ----------------------------------------------------------

if os.path.exists(excel_file):

    wb = load_workbook(excel_file)

    if "Benchmark Results" in wb.sheetnames:
        ws = wb["Benchmark Results"]

        # Remove old data
        ws.delete_rows(2, ws.max_row)

    else:
        ws = wb.create_sheet("Benchmark Results")

else:

    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark Results"

# ----------------------------------------------------------
# Header
# ----------------------------------------------------------

headers = [
    "Repository",
    "Retriever",
    "Top K",
    "Concurrent Queries",
    "Successful Queries",
    "Failed Queries",
    "Total Time (s)",
    "Average Latency (ms)",
    "Minimum Latency (ms)",
    "Maximum Latency (ms)",
    "Throughput (queries/sec)",
]

if ws.max_row == 1 and ws["A1"].value is None:

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(row=1, column=col)

        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            fill_type="solid",
            start_color="4F81BD",
            end_color="4F81BD",
        )
        cell.alignment = Alignment(horizontal="center")

# ----------------------------------------------------------
# Write Benchmark Results
# ----------------------------------------------------------

for row in results:

    (
        concurrent_queries,
        successful,
        failed,
        total_time,
        avg_latency,
        min_latency,
        max_latency,
        throughput,
    ) = row

    ws.append([
        REPOSITORY_NAME,
        RETRIEVER_NAME,
        TOP_K,
        concurrent_queries,
        successful,
        failed,
        round(total_time, 2),
        round(avg_latency, 2),
        round(min_latency, 2),
        round(max_latency, 2),
        round(throughput, 2),
    ])

# ----------------------------------------------------------
# Auto-size Columns
# ----------------------------------------------------------

for column_cells in ws.columns:

    max_length = max(
        len(str(cell.value)) if cell.value else 0
        for cell in column_cells
    )

    ws.column_dimensions[
        get_column_letter(column_cells[0].column)
    ].width = max_length + 4

# ----------------------------------------------------------
# Freeze Header
# ----------------------------------------------------------

ws.freeze_panes = "A2"

# ----------------------------------------------------------
# Save Workbook
# ----------------------------------------------------------

wb.save(excel_file)

print()
print("=" * 100)
print(f"Benchmark results updated successfully: {excel_file}")
print("=" * 100)